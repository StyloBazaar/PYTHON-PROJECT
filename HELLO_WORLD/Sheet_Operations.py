import openpyxl as xl
from openpyxl.chart import BarChart, Reference
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
# print(sheet.max_row)
for row in range(2, sheet.max_row+1):
    price = sheet.cell(row, 3).value
    discounted_price = price * 0.9
    discounted_price_cell = sheet.cell(row, 4)
    discounted_price_cell.value = discounted_price

values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'a5')
wb.save('transactions2.xlsx')

